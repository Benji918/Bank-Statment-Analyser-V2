import io
import json
import logging
from typing import Any, Dict

logger = logging.getLogger(__name__)


def generate_pdf_report(insight_data: Dict[str, Any]) -> bytes:
    """Generate a styled PDF report from insight JSON using ReportLab."""
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle
    from reportlab.lib import colors

    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4)
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("Bank Statement Analysis Report", styles["h1"]))
    story.append(Spacer(1, 12))

    summary_data = [
        ["Metric", "Value"],
        ["Total Income", f"£{insight_data.get('total_income', 0):,.2f}"],
        ["Total Expenses", f"£{insight_data.get('total_expenses', 0):,.2f}"],
        ["Net Balance", f"£{insight_data.get('net_balance', 0):,.2f}"],
        ["Savings Rate", f"{insight_data.get('savings_rate_percent', 0):.1f}%"],
    ]
    table = Table(summary_data)
    table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0000EE")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("GRID", (0, 0), (-1, -1), 0.5, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 10),
    ]))
    story.append(table)
    story.append(Spacer(1, 12))

    insights_list = insight_data.get("actionable_insights", [])
    if insights_list:
        story.append(Paragraph("Actionable Insights", styles["h2"]))
        for insight in insights_list:
            story.append(Paragraph(f"• {insight}", styles["Normal"]))
        story.append(Spacer(1, 12))

    doc.build(story)
    return buffer.getvalue()


def generate_excel(insight_data: Dict[str, Any]) -> bytes:
    """Generate a multi-sheet Excel workbook from insight JSON using openpyxl."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()

    # Summary sheet
    ws_summary = wb.active
    ws_summary.title = "Summary"
    headers = ["Metric", "Value"]
    ws_summary.append(headers)
    ws_summary.append(["Total Income", insight_data.get("total_income", 0)])
    ws_summary.append(["Total Expenses", insight_data.get("total_expenses", 0)])
    ws_summary.append(["Net Balance", insight_data.get("net_balance", 0)])
    ws_summary.append(["Savings Rate (%)", insight_data.get("savings_rate_percent", 0)])

    # Categories sheet
    ws_cats = wb.create_sheet("Spending By Category")
    ws_cats.append(["Category", "Amount (£)"])
    for cat, amt in insight_data.get("spending_by_category", {}).items():
        ws_cats.append([cat, amt])

    # Recurring transactions sheet
    ws_recur = wb.create_sheet("Recurring Transactions")
    ws_recur.append(["Type", "Description", "Amount (£)", "Frequency"])
    for t in insight_data.get("recurring_debits", []):
        ws_recur.append(["Debit", t.get("description"), t.get("amount"), t.get("frequency")])
    for t in insight_data.get("recurring_credits", []):
        ws_recur.append(["Credit", t.get("description"), t.get("amount"), t.get("frequency")])

    buffer = io.BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


def generate_json(insight_data: Dict[str, Any]) -> str:
    """Serialise insight data to JSON string."""
    return json.dumps(insight_data, indent=2)
